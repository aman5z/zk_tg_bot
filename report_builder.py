"""
report_builder.py
Build attendance reports in XLSX, PNG, and PDF formats.
Supports department-priority sorting and multiple visual templates.
"""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import matplotlib
matplotlib.use('Agg')  # non-interactive backend — must be before pyplot import
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# ─── Department sort order ────────────────────────────────────────────────────
# TEACHING first, then ADMIN, SUPPORT, DRIVER, CLEANING STAFF, then rest alpha.
DEPT_ORDER = [
    'TEACHING',
    'ADMIN',
    'SUPPORT',
    'DRIVER',
    'CLEANING STAFF',
]


def dept_sort_key(dept: str) -> tuple:
    upper = (dept or '').upper().strip()
    try:
        return (DEPT_ORDER.index(upper), dept)
    except ValueError:
        return (len(DEPT_ORDER), dept)


# ─── Visual templates ─────────────────────────────────────────────────────────
TEMPLATES: Dict[str, Dict] = {
    'default': {
        'label':       'Default (Blue)',
        'header_bg':   '#2196F3',
        'header_fg':   'white',
        'row_even':    '#E3F2FD',
        'row_odd':     '#FFFFFF',
        'border':      '#BBDEFB',
    },
    'dark': {
        'label':       'Dark (Navy)',
        'header_bg':   '#1A237E',
        'header_fg':   'white',
        'row_even':    '#E8EAF6',
        'row_odd':     '#FFFFFF',
        'border':      '#C5CAE9',
    },
    'green': {
        'label':       'Green (School)',
        'header_bg':   '#1B5E20',
        'header_fg':   'white',
        'row_even':    '#E8F5E9',
        'row_odd':     '#FFFFFF',
        'border':      '#A5D6A7',
    },
}

_COL_WIDTHS_XLS = {'A': 10, 'B': 36, 'C': 22}   # column letter → character width
_COL_WIDTHS_FIG = [0.09, 0.42, 0.22]             # relative widths for figure table


# ─── Internal helpers ─────────────────────────────────────────────────────────

def _hex_to_rgb(h: str) -> Tuple[float, float, float]:
    h = h.lstrip('#')
    if len(h) != 6 or not all(c in '0123456789abcdefABCDEF' for c in h):
        raise ValueError(f"Invalid hex colour: '#{h}'")
    return tuple(int(h[i:i+2], 16) / 255.0 for i in (0, 2, 4))


def _sort_and_filter(
    absent: List[Dict],
    departments: str = 'ALL',
    extra_exclude_badges: Optional[set] = None,
) -> List[Dict]:
    """Filter by department and sort by DEPT_ORDER, then name."""
    result = absent
    if departments and departments.strip().upper() != 'ALL':
        selected = {d.strip().upper() for d in departments.split(',') if d.strip()}
        result = [e for e in result if (e.get('dept') or '').upper() in selected]
    if extra_exclude_badges:
        result = [e for e in result if e.get('badge') not in extra_exclude_badges]
    result = sorted(result, key=lambda e: (dept_sort_key(e.get('dept', '')),
                                           (e.get('name') or '').upper()))
    return result


def _to_rows(emp_list: List[Dict]) -> List[Dict]:
    return [{'Badge': e['badge'], 'Name': e['name'], 'Department': e['dept']}
            for e in emp_list]


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def build_xlsx(rows: List[Dict], title: str = '', template: str = 'default',
               subtitle: str = '') -> BytesIO:
    tpl = TEMPLATES.get(template, TEMPLATES['default'])
    df = pd.DataFrame(rows, columns=['Badge', 'Name', 'Department'])
    buf = BytesIO()

    # Determine how many header rows precede the column header
    extra_rows = (1 if title else 0) + (1 if subtitle else 0)

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Absent', startrow=extra_rows)
        ws = writer.sheets['Absent']

        header_row = extra_rows + 1  # 1-indexed row where Badge/Name/Dept header lives

        # Optional title row
        if title:
            ws.merge_cells('A1:C1')
            tc = ws['A1']
            tc.value = title
            tc.font = Font(bold=True, size=13, color='FFFFFF')
            tc.fill = PatternFill(
                start_color=tpl['header_bg'].lstrip('#'),
                end_color=tpl['header_bg'].lstrip('#'),
                fill_type='solid')
            tc.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 22

        # Optional subtitle row (e.g. "Report generated on: …")
        if subtitle:
            subtitle_row = 2 if title else 1
            ws.merge_cells(f'A{subtitle_row}:C{subtitle_row}')
            sc = ws[f'A{subtitle_row}']
            sc.value = subtitle
            sc.font = Font(italic=True, size=10, color='444444')
            sc.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[subtitle_row].height = 16

        # Style the Badge / Name / Department header row
        hdr_fill = PatternFill(
            start_color=tpl['header_bg'].lstrip('#'),
            end_color=tpl['header_bg'].lstrip('#'),
            fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', size=11)
        for col in range(1, 4):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')

        # Alternating row fill
        even_fill = PatternFill(
            start_color=tpl['row_even'].lstrip('#'),
            end_color=tpl['row_even'].lstrip('#'),
            fill_type='solid')
        data_start = header_row + 1
        for row_idx in range(data_start, data_start + len(rows)):
            if (row_idx - data_start) % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).fill = even_fill

        # Column widths
        for col_letter, width in _COL_WIDTHS_XLS.items():
            ws.column_dimensions[col_letter].width = width

    buf.seek(0)
    return buf


def build_timings_xlsx(
    rows: List[Dict],
    report_date: date,
    departments_label: str = 'ALL',
    mode: str = 'summary',
    template: str = 'default',
    dept_totals: Optional[Dict[str, Dict[str, int]]] = None,
    generated_at: Optional[str] = None,
) -> BytesIO:
    """Build grouped timings XLSX report."""
    tpl = TEMPLATES.get(template, TEMPLATES['default'])
    if generated_at is None:
        generated_at = datetime.now().strftime('%d-%m-%Y %I:%M%p')
    mode = (mode or 'summary').strip().lower()
    if mode not in {'summary', 'all'}:
        mode = 'summary'

    ordered_rows = sorted(
        rows,
        key=lambda r: (dept_sort_key(r.get('Department', '')),
                       (r.get('Employee Name') or '').upper())
    )
    by_dept: Dict[str, List[Dict]] = {}
    for row in ordered_rows:
        by_dept.setdefault(row.get('Department') or 'Unknown', []).append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Timings'

    ws.merge_cells('A1:G1')
    ws['A1'] = f"Timings Report — {report_date.strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill(
        start_color=tpl['header_bg'].lstrip('#'),
        end_color=tpl['header_bg'].lstrip('#'),
        fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Departments: {departments_label}  |  Generated: {generated_at}"
    ws['A2'].font = Font(italic=True, size=10, color='444444')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:G3')
    ws['A3'] = f"View: {'Summary (In/Out)' if mode == 'summary' else 'All Punches'}"
    ws['A3'].font = Font(italic=True, size=10, color='444444')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        'Employee Name', 'Badge', 'Department'
    ]
    if mode == 'summary':
        headers += ['Check-In', 'Check-Out', 'Total Hours', 'Status']
    else:
        headers += ['Punch #', 'Time', 'Device/Sensor', 'Total Punches']

    header_fill = PatternFill(
        start_color=tpl['header_bg'].lstrip('#'),
        end_color=tpl['header_bg'].lstrip('#'),
        fill_type='solid')
    for col_idx, head in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=head)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')

    even_fill = PatternFill(
        start_color=tpl['row_even'].lstrip('#'),
        end_color=tpl['row_even'].lstrip('#'),
        fill_type='solid')
    dept_fill = PatternFill(
        start_color=tpl['border'].lstrip('#'),
        end_color=tpl['border'].lstrip('#'),
        fill_type='solid')

    row_idx = 6
    stripe_idx = 0
    for dept in sorted(by_dept, key=dept_sort_key):
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        dcell = ws.cell(row=row_idx, column=1, value=f'Department: {dept}')
        dcell.font = Font(bold=True, color='333333')
        dcell.fill = dept_fill
        dcell.alignment = Alignment(horizontal='left')
        row_idx += 1

        dept_rows = by_dept[dept]
        for entry in dept_rows:
            values = [entry.get('Employee Name', ''), entry.get('Badge', ''), entry.get('Department', '')]
            if mode == 'summary':
                values += [
                    entry.get('Check-In', '—'),
                    entry.get('Check-Out', '—'),
                    entry.get('Total Hours', ''),
                    entry.get('Status', ''),
                ]
            else:
                values += [
                    entry.get('Punch #', '—'),
                    entry.get('Time', '—'),
                    entry.get('Device/Sensor', '—'),
                    entry.get('Total Punches', 0),
                ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if stripe_idx % 2 == 0:
                    cell.fill = even_fill
                if col_idx >= 4:
                    cell.alignment = Alignment(horizontal='center')
            row_idx += 1
            stripe_idx += 1

        present_count = absent_count = complete_count = in_only_count = 0
        if mode == 'summary':
            complete_count = sum(1 for r in dept_rows if r.get('Status') == '✅ Complete')
            in_only_count = sum(1 for r in dept_rows if r.get('Status') == '⏳ In Only')
            absent_count = sum(1 for r in dept_rows if r.get('Status') == '❌ Absent')
            present_count = complete_count + in_only_count
            if dept_totals and dept in dept_totals:
                present_count = int(dept_totals[dept].get('present', present_count))
                absent_count = int(dept_totals[dept].get('absent', absent_count))
        else:
            present_count = len({r.get('Badge') for r in dept_rows if r.get('Total Punches', 0)})
            absent_count = 0
            if dept_totals and dept in dept_totals:
                absent_count = int(dept_totals[dept].get('absent', 0))
                present_count = int(dept_totals[dept].get('present', present_count))

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        if mode == 'summary':
            subtotal_txt = (
                f'Subtotal — ✅ Complete: {complete_count} | '
                f'⏳ In Only: {in_only_count} | ❌ Absent: {absent_count}'
            )
        else:
            subtotal_txt = (
                f'Subtotal — Present Staff: {present_count} | '
                f'Absent Staff: {absent_count}'
            )
        scell = ws.cell(row=row_idx, column=1, value=subtotal_txt)
        scell.font = Font(bold=True, italic=True, color='333333')
        scell.fill = dept_fill
        ws.cell(row=row_idx, column=7, value=len(dept_rows)).fill = dept_fill
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='center')
        row_idx += 2

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.freeze_panes = 'A6'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _timings_table_fig(
    rows: List[List[str]],
    headers: List[str],
    title: str,
    template: str = 'default',
) -> plt.Figure:
    tpl = TEMPLATES.get(template, TEMPLATES['default'])
    n = len(rows)
    fig_h = max(3.2, 1.4 + n * 0.34)
    fig, ax = plt.subplots(figsize=(14, fig_h))
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    ax.axis('off')
    if rows:
        cell_colors = [
            [tpl['row_even'] if i % 2 == 0 else tpl['row_odd']] * len(headers)
            for i in range(n)
        ]
        col_colors = [_hex_to_rgb(tpl['header_bg'])] * len(headers)
        tbl = ax.table(
            cellText=rows,
            colLabels=headers,
            cellColours=cell_colors,
            colColours=col_colors,
            loc='upper center',
            cellLoc='left',
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        for j in range(len(headers)):
            tbl[0, j].set_text_props(color=tpl['header_fg'], fontweight='bold')
            width = 1.0 / len(headers)
            for i in range(n + 1):
                tbl[i, j].set_width(width)
    else:
        ax.text(0.5, 0.5, 'No rows found.', ha='center', va='center', fontsize=14, transform=ax.transAxes)
    fig.suptitle(title, fontsize=10, fontweight='bold', y=0.99, color='#333333')
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


def build_timings_report(
    rows: List[Dict],
    report_date: date,
    departments_label: str = 'ALL',
    mode: str = 'summary',
    fmt: str = 'xlsx',
    template: str = 'default',
    dept_totals: Optional[Dict[str, Dict[str, int]]] = None,
    generated_at: Optional[str] = None,
) -> BytesIO:
    mode = (mode or 'summary').strip().lower()
    if mode not in {'summary', 'all'}:
        mode = 'summary'
    fmt = (fmt or 'xlsx').strip().lower()
    if generated_at is None:
        generated_at = datetime.now().strftime('%d-%m-%Y %I:%M%p')

    if fmt == 'xlsx':
        return build_timings_xlsx(
            rows=rows,
            report_date=report_date,
            departments_label=departments_label,
            mode=mode,
            template=template,
            dept_totals=dept_totals,
            generated_at=generated_at,
        )

    headers = ['Employee Name', 'Badge', 'Department']
    if mode == 'summary':
        headers += ['Check-In', 'Check-Out', 'Total Hours', 'Status']
    else:
        headers += ['Punch #', 'Time', 'Device/Sensor', 'Total Punches']
    ordered = sorted(
        rows,
        key=lambda r: (dept_sort_key(r.get('Department', '')), (r.get('Employee Name') or '').upper())
    )
    def _display_value(header: str, value) -> str:
        text = str(value if value is not None else '')
        if header == 'Status':
            return (text
                    .replace('✅ Complete', 'Complete')
                    .replace('⏳ In Only', 'In Only')
                    .replace('❌ Absent', 'Absent'))
        return text

    table_rows = [[_display_value(h, r.get(h, '')) for h in headers] for r in ordered]
    title = (
        f"Timings Report — {report_date.strftime('%d/%m/%Y')} | "
        f"Departments: {departments_label} | Generated: {generated_at} | "
        f"View: {'Summary (In/Out)' if mode == 'summary' else 'All Punches'}"
    )

    if fmt == 'png':
        fig = _timings_table_fig(table_rows, headers, title=title, template=template)
        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor='white', edgecolor='none')
        plt.close(fig)
        buf.seek(0)
        return buf

    if fmt == 'pdf':
        buf = BytesIO()
        with PdfPages(buf) as pdf:
            page_size = 36
            pages = [table_rows[i:i + page_size] for i in range(0, len(table_rows), page_size)] or [[]]
            for idx, page_rows in enumerate(pages, 1):
                page_title = title
                if len(pages) > 1:
                    page_title = f"{title} (Page {idx}/{len(pages)})"
                fig = _timings_table_fig(page_rows, headers, title=page_title, template=template)
                pdf.savefig(fig, bbox_inches='tight', facecolor='white', edgecolor='none')
                plt.close(fig)
        buf.seek(0)
        return buf

    raise ValueError(f'Unsupported timings format: {fmt}')


# ─── PNG ──────────────────────────────────────────────────────────────────────

def build_png(rows: List[Dict], title: str = '', template: str = 'default',
              subtitle: str = '') -> BytesIO:
    tpl = TEMPLATES.get(template, TEMPLATES['default'])
    n = len(rows)

    fig_h = max(2.5, 0.95 + n * 0.38)
    fig, ax = plt.subplots(figsize=(10, fig_h))
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')
    ax.axis('off')

    if not rows:
        ax.text(0.5, 0.5, 'No absences today — All present!', ha='center', va='center',
                fontsize=14, transform=ax.transAxes)
    else:
        col_labels = ['Badge', 'Name', 'Department']
        cell_data = [[r.get('Badge', ''), r.get('Name', ''), r.get('Department', '')]
                     for r in rows]
        cell_colors = [
            [tpl['row_even'] if i % 2 == 0 else tpl['row_odd']] * 3
            for i in range(n)
        ]
        col_colors = [_hex_to_rgb(tpl['header_bg'])] * 3

        tbl = ax.table(
            cellText=cell_data,
            colLabels=col_labels,
            cellColours=cell_colors,
            colColours=col_colors,
            loc='upper center',
            cellLoc='left',
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(9)

        # Style header text
        for j in range(3):
            tbl[0, j].set_text_props(color=tpl['header_fg'], fontweight='bold')

        # Column widths
        for j, w in enumerate(_COL_WIDTHS_FIG):
            for i in range(n + 1):
                tbl[i, j].set_width(w)

    full_title = title
    if subtitle:
        full_title = f"{title}\n{subtitle}" if title else subtitle
    if full_title:
        fig.suptitle(full_title, fontsize=11, fontweight='bold', y=0.99,
                     color='#333333')

    plt.tight_layout(rect=[0, 0, 1, 0.96 if full_title else 1.0])
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    return buf


# ─── PDF ──────────────────────────────────────────────────────────────────────

_ROWS_PER_PAGE = 48


def build_pdf(rows: List[Dict], title: str = '', template: str = 'default',
              subtitle: str = '') -> BytesIO:
    tpl = TEMPLATES.get(template, TEMPLATES['default'])
    full_title = title
    if subtitle:
        full_title = f"{title}\n{subtitle}" if title else subtitle
    # When rows is empty produce a single empty page; otherwise paginate.
    if rows:
        pages = [rows[i:i + _ROWS_PER_PAGE]
                 for i in range(0, len(rows), _ROWS_PER_PAGE)]
    else:
        pages = [[]]

    buf = BytesIO()
    with PdfPages(buf) as pdf:
        for page_idx, page_rows in enumerate(pages):
            n = len(page_rows)
            fig_h = max(3.0, 1.3 + n * 0.38)
            fig, ax = plt.subplots(figsize=(10, fig_h))
            ax.set_facecolor('white')
            fig.patch.set_facecolor('white')
            ax.axis('off')

            page_title = full_title
            if len(pages) > 1:
                page_title += f'  (page {page_idx + 1}/{len(pages)})'

            if page_rows:
                col_labels = ['Badge', 'Name', 'Department']
                cell_data = [[r.get('Badge', ''), r.get('Name', ''), r.get('Department', '')]
                             for r in page_rows]
                cell_colors = [
                    [tpl['row_even'] if i % 2 == 0 else tpl['row_odd']] * 3
                    for i in range(n)
                ]
                col_colors = [_hex_to_rgb(tpl['header_bg'])] * 3

                tbl = ax.table(
                    cellText=cell_data,
                    colLabels=col_labels,
                    cellColours=cell_colors,
                    colColours=col_colors,
                    loc='upper center',
                    cellLoc='left',
                )
                tbl.auto_set_font_size(False)
                tbl.set_fontsize(9)
                for j in range(3):
                    tbl[0, j].set_text_props(color=tpl['header_fg'], fontweight='bold')
                for j, w in enumerate(_COL_WIDTHS_FIG):
                    for i in range(n + 1):
                        tbl[i, j].set_width(w)
            else:
                ax.text(0.5, 0.5, 'No absences today — All present!', ha='center', va='center',
                        fontsize=14, transform=ax.transAxes)

            if page_title:
                fig.suptitle(page_title, fontsize=11, fontweight='bold',
                             y=0.99, color='#333333')

            plt.tight_layout(rect=[0, 0, 1, 0.96 if page_title else 1.0])
            pdf.savefig(fig, bbox_inches='tight', facecolor='white', edgecolor='none')
            plt.close(fig)

    buf.seek(0)
    return buf


# ─── Main entry point ─────────────────────────────────────────────────────────

def build_absent_report(
    absent: List[Dict],
    report_date: date,
    departments: str = 'ALL',
    formats: str = 'xlsx',
    template: str = 'default',
    extra_exclude_badges: Optional[set] = None,
    generated_at: Optional[str] = None,
) -> Dict[str, Tuple[BytesIO, str]]:
    """
    Build absent report in the requested format(s).

    Returns dict of  format_key → (BytesIO, filename).
    format_key values: 'xlsx', 'png', 'pdf'
    generated_at: optional pre-formatted timestamp string; defaults to now.
    """
    emp_list = _sort_and_filter(absent, departments, extra_exclude_badges)
    rows = _to_rows(emp_list)

    if generated_at is None:
        now = datetime.now()
        generated_at = now.strftime('%d-%m-%Y %I:%M%p')

    date_label = report_date.strftime('%d %B %Y')
    date_file  = report_date.strftime('%d-%m-%Y')
    title = f"Absent Report — {date_label}"
    subtitle = f"Report generated on: {generated_at}"
    filename_base = f"{date_file} Attendance Report"

    fmt_set = {f.strip().lower() for f in formats.split(',')}
    send_all = 'all' in fmt_set

    result: Dict[str, Tuple[BytesIO, str]] = {}

    if send_all or 'xlsx' in fmt_set:
        result['xlsx'] = (build_xlsx(rows, title, template, subtitle=subtitle),
                          f'{filename_base}.xlsx')
    if send_all or 'png' in fmt_set:
        result['png'] = (build_png(rows, title, template, subtitle=subtitle),
                         f'{filename_base}.png')
    if send_all or 'pdf' in fmt_set:
        result['pdf'] = (build_pdf(rows, title, template, subtitle=subtitle),
                         f'{filename_base}.pdf')

    return result
